from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
# from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from init_bot import bot
import os
import openpyxl
from datetime import datetime, date, timedelta

from DBController import DBController


class FSMDownload(StatesGroup):
    parse_start = State()


async def download(message: types.Message) -> None:
    if DBController().getEntryByChatID(message.chat.id)["role"] == 1:
        await FSMDownload.parse_start.set()
        await message.answer(text='Скиньте документ с расписанием в формате xlsx\nДля отменны действия введите "отмена"')
        await message.delete()
    else:
        await message.answer('У вас недостаточно прав')


async def cancelDownload(message: types.Message, state: FSMContext) -> None:
    cur_state = await state.get_state()
    if cur_state is None:
        return
    await message.answer(text="Операция успешно отменена")
    await state.finish()


def deploy(groups: list, start_row: int, sheet: openpyxl.Workbook,
           week: list[str], time: list[str]) -> None:
    """_summary_

    Args:
        groups (list): _description_
        start_row (int): _description_
        sheet (openpyxl.Workbook): _description_
        week (list[str]): _description_
        time (list[str]): _description_
    """
    for group in groups:
        begin = start_row
        for day in week:
            for number_of_obj in time:
                subject = sheet.cell(row=begin, column=group['id'] + 1).value
                room = sheet.cell(row=begin, column=group['id'] + 2).value
                educator = sheet.cell(row=begin, column=group['id'] + 3).value
                if educator:
                    print(f'<{day} -- {number_of_obj} -- {subject} -- {room} -- {group["name"]} -- {educator}>')

                    if not DBController().getUser(educator):
                        DBController().addUser(educator)

                    if not DBController().getGroup(group["name"]):
                        DBController().addGroup(group["name"])

                    if not DBController().getSubject(subject):
                        DBController().addSubject(subject)

                    if not DBController().getScheduleEntry(educator, group['name'], subject, number_of_obj, room, day):
                        DBController().addScheduleEntry(educator, group['name'], subject, number_of_obj, room, day)

                begin += 1


async def parse(message: types.Message, file_name: str) -> None:
    """_summary_

    Args:
        message (types.Message): _description_
        file_name (str): _description_
    """
    try:
        wb: openpyxl.Workbook = openpyxl.load_workbook(filename=file_name, data_only=True)
        for sheetname in wb.sheetnames:
            sheet: openpyxl.Workbook = wb[sheetname]
            i: int = 0
            start_row: int = 0
            groups: list[dict] = []
            for row in sheet.values:
                i += 1
                if isinstance(row[0], str):
                    if row[0].lower().strip() == 'дни':
                        groups: list[dict] = [{'name': row[i], 'id': i} for i in range(3, len(row)) if row[i] is not None]
                        start_row: int = i + 2
                        break

            time: list[str] = ['9-00', '10-45', '13-00', '14-45', '16-25', '18-15', '19-45']
            date_row: int = start_row
            dates = sheet[f'A{date_row}':f'A{date_row + 6}']
            start_date = None
            for cell in dates:
                if isinstance(cell[0].value, datetime):
                    start_date = cell[0].value
                    break
            start_date: date = start_date.date()
            week: list[str] = [(start_date + timedelta(days=i)).strftime('%d.%m.%y') for i in range(0, 6)]
            deploy(groups, start_row, sheet, week, time)

        await message.answer("Данные успешно добавлены")
        # await message.delete()
    except:
        await message.answer("Не удалось обработать документ")


async def getFile(message: types.Message, state: FSMContext) -> None:
    """_summary_

    Args:
        message (types.Message): _description_
    """

    try:
        await message.answer("Начинаю загрузку документа")
        save_dir = "./"
        file_name = message.document.file_name
        await message.document.download(destination_file=save_dir + file_name)
        await message.answer("Документ загружен, начинаю обработку")
        await parse(message, file_name)
    except:
        await message.answer("Произошла непредвиденная ошибка")
    finally:
        await state.finish()


def Admin_Handlers(dp: Dispatcher):
    dp.register_message_handler(download, commands=['download'], state=None)
    dp.register_message_handler(cancelDownload, state="parse_start", commands="отмена")
    dp.register_message_handler(cancelDownload, Text(equals="отмена", ignore_case=True), state="parse_start")
    dp.register_message_handler(getFile, content_types=['document'], state=FSMDownload.parse_start)
