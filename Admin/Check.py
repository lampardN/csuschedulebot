from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
# from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from init_bot import bot
import os
import openpyxl
from datetime import datetime, date, timedelta, time

from DBController import DBController


class Entry:
    def __init__(self, educator, custom_time, room, custom_date):
        self.educator = educator
        self.custom_time = custom_time
        self.room = room
        self.custom_date = custom_date
        self.count = 1

    def __str__(self):
        return f"{self.educator}\n{self.custom_time}\n{self.room}\n{self.custom_date}\n"

    def __repr__(self):
        return repr([self.educator, self.custom_time, self.room, self.custom_date, self.count])

    def __eq__(self, other):
        if not isinstance(other, Entry):
            raise TypeError("other type is not Entry")
        if self.educator == other.educator and self.custom_time == other.custom_time and self.room == other.room and self.custom_date == other.custom_date:
            self.count += 1
            return True
        return False


class FSMCheck(StatesGroup):
    parse_start = State()


async def download(message: types.Message) -> None:
    if DBController().getEntryByChatID(message.chat.id)["role"] == 1:
        await FSMCheck.parse_start.set()
        await message.answer(
            text='Скиньте документ с расписанием в формате xlsx\nДля отменны действия введите "отмена"')
        await message.delete()
    else:
        await message.answer('У вас недостаточно прав')


async def cancelDownload(message: types.Message, state: FSMContext) -> None:
    cur_state = await state.get_state()
    if cur_state is None:
        return
    await message.answer(text="Операция успешно отменена")
    await state.finish()


def deploy(groups: list, start_row: int, sheet: openpyxl.Workbook,
           week: list[date], custom_time: list[time], entries: list[Entry]) -> None:
    """_summary_

    Args:
        groups (list): _description_
        start_row (int): _description_
        sheet (openpyxl.Workbook): _description_
        week (list[str]): _description_
        custom_time (list[str]): _description_
        :param groups:
        :param start_row:
        :param sheet:
        :param week:
        :param custom_time:
        :param entries:
    """

    for group in groups:
        begin = start_row
        for day in week:
            for number_of_obj in custom_time:
                room = sheet.cell(row=begin, column=group['id'] + 2).value
                if isinstance(room, int):
                    room = str(room)
                educator = sheet.cell(row=begin, column=group['id'] + 3).value
                if educator:
                    flag = 0
                    new_entry = Entry(educator, number_of_obj, room, day)
                    for entry in entries:
                        if entry == new_entry:
                            flag = 1
                    if not flag or not entries:
                        entries.append(new_entry)
                begin += 1


async def parse(message: types.Message, file_name: str) -> None:
    """_summary_

    Args:
        message (types.Message): _description_
        file_name (str): _description_
    """
    try:
        entries: list[Entry] = []
        wb: openpyxl.Workbook = openpyxl.load_workbook(filename=file_name, data_only=True)
        for sheetname in wb.sheetnames:
            sheet: openpyxl.Workbook = wb[sheetname]
            i: int = 0
            start_row: int = 0
            groups: list[dict] = []
            for row in sheet.values:
                i += 1
                if isinstance(row[0], str):
                    if row[0].lower().strip() == 'дни':
                        groups: list[dict] = [{'name': row[i], 'id': i} for i in range(3, len(row)) if
                                              row[i] is not None]
                        start_row: int = i + 2
                        break

            custom_time: list[time] = [time(9, 0),
                                       time(10, 45),
                                       time(13, 0),
                                       time(14, 45),
                                       time(16, 25),
                                       time(18, 25),
                                       time(19, 45)]
            date_row: int = start_row
            dates = sheet[f'A{date_row}':f'A{date_row + 6}']
            start_date = None
            for cell in dates:
                if isinstance(cell[0].value, datetime):
                    start_date = cell[0].value
                    break
            week: list[date] = [(start_date + timedelta(days=i)) for i in range(0, 6)]
            deploy(groups, start_row, sheet, week, custom_time, entries)

        await message.answer("Данные успешно добавлены")
        await check(message, entries)
        # await message.delete()
    except:
        await message.answer("Не удалось обработать документ")


async def check(message: types.Message, entries: list[Entry]):
    from operator import attrgetter
    from pprint import pprint

    entries.sort(key=attrgetter('custom_date', 'custom_time'))
    # pprint(entries)
    unique_counter = {}
    for entry in entries:
        if entry.custom_date not in unique_counter.keys():
            unique_counter[entry.custom_date] = {}
        if entry.custom_time not in unique_counter[entry.custom_date].keys():
            unique_counter[entry.custom_date][entry.custom_time] = {}
        if entry.room not in unique_counter[entry.custom_date][entry.custom_time].keys():
            unique_counter[entry.custom_date][entry.custom_time][entry.room] = []
        if entry.educator not in unique_counter[entry.custom_date][entry.custom_time][entry.room]:
            unique_counter[entry.custom_date][entry.custom_time][entry.room].append(entry.educator)

    # pprint(unique_counter)
    for custom_date in unique_counter.keys():
        new_message = f"{custom_date.strftime('%d.%m.%y')}\n"
        for custom_time in unique_counter[custom_date].keys():
            free_rooms: list[str] = ['101', '102', '105', '203', '204', '205', '206', '301', '303', '304', '305', '306',
                                     '307', '210', '211', '220', '222', '224', '232', '309', '310', '311', 'спорт.зал']
            true_rooms = ""
            false_rooms = ""
            for room in unique_counter[custom_date][custom_time].keys():
                if len(unique_counter[custom_date][custom_time][room]) == 1:
                    true_rooms += f'{room} {unique_counter[custom_date][custom_time][room]} || '
                else:
                    false_rooms += f'{room} {unique_counter[custom_date][custom_time][room]} || '
                if room in free_rooms:
                    free_rooms.pop(free_rooms.index(room))
            true_rooms = true_rooms[:-4]
            false_rooms = false_rooms[:-4]
            new_message += f"<{custom_time.isoformat('minutes')}>\n\n"
            new_message += f"Правильные аудитории: {true_rooms}\n\n"
            new_message += f"Ошибки: {false_rooms}\n\n"
            new_message += f"Свободные аудитории: {', '.join(el for el in free_rooms)}\n\n"
        new_message = new_message[:-2]
        await message.answer(new_message)
    # await message.answer('Проверка пройдена')


async def getFile(message: types.Message, state: FSMContext) -> None:
    """_summary_

    Args:
        message (types.Message): _description_
        :param message:
        :param state:
    """

    try:
        await message.answer("Начинаю загрузку документа")
        save_dir = "./"
        file_name = message.document.file_name
        await message.document.download(destination_file=save_dir + file_name)
        await message.answer("Документ загружен, начинаю обработку")
        await parse(message, file_name)
    except:
        await message.answer("Произошла непредвиденная ошибка")
    finally:
        await state.finish()


def Check_Handlers(dp: Dispatcher):
    dp.register_message_handler(download, commands=['check'], state=None)
    dp.register_message_handler(cancelDownload, state="parse_start", commands="отмена")
    dp.register_message_handler(cancelDownload, Text(equals="отмена", ignore_case=True), state="parse_start")
    dp.register_message_handler(getFile, content_types=['document'], state=FSMCheck.parse_start)
